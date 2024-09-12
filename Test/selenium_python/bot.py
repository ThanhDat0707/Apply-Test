import os
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

chrome_options = Options()
chrome_options.add_argument("--incognito")
chrome_options.add_argument("--window-size=1920x1080")
chrome_options.add_argument("--ignore-certificate-errors")   # Bỏ qua lỗi SSL
chrome_options.add_argument("--allow-insecure-localhost")    # Cho phép localhost không an toàn
chrome_options.add_argument("--disable-web-security")        # Vô hiệu hóa bảo mật web 

s = Service("D:/Test/chromedriver-win64/chromedriver.exe")
driver = webdriver.Chrome(service=s, options=chrome_options)

urls = [
    "https://news.ycombinator.com/",
    "https://webdriveruniversity.com/",
    "https://www.saucedemo.com/",
    "https://ultimateqa.com/automation",
    "https://github.com/nadvolod/react-shopping-cart",
    "https://www.telerik.com/support/demos",
    "https://demoqa.com/",
    "https://compendiumdev.co.uk/",
    "https://demo.applitools.com/",
    "http://practice.automationtesting.in/"
]

titles = []
contents = []
html_sources = []

for url in urls:
    driver.get(url)
    
    titles.append(driver.title)
    
    try:
        paragraphs = driver.find_elements(By.TAG_NAME, "p")
        spans = driver.find_elements(By.TAG_NAME, "span")
        h1s = driver.find_elements(By.TAG_NAME, "h1")
        h2s = driver.find_elements(By.TAG_NAME, "h2")
        h3s = driver.find_elements(By.TAG_NAME, "h3")
        h4s = driver.find_elements(By.TAG_NAME, "h4")
        br = driver.find_elements(By.TAG_NAME, "br")
        page_source = driver.page_source

        content = " ".join([p.text for p in paragraphs + spans + h1s + h2s + h3s + h4s + br])
        
        contents.append(content if content else "No content found")
        html_sources.append(page_source if page_source else "No HTML sources found")
    except Exception as e:
        contents.append(f"Error retrieving content: {str(e)}")
        html_sources.append(f"Error retrieving content: {str(e)}")

driver.quit()

df = pd.DataFrame({
    "URL": urls,
    "Title": titles,
    "Content": contents,
    "HTML Sources": html_sources
})

output_file = "page_titles_with_content_HTML_source.xlsx"

if os.path.exists(output_file):
    os.remove(output_file)
    print(f"Deleted old file: {output_file}")

df.to_excel(output_file, index=False)

wb = load_workbook(output_file)
ws = wb.active

for column_cells in ws.columns:
    max_length = 0
    column = column_cells[0].column_letter  
    for cell in column_cells:
        try:
            max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    adjusted_width = max_length + 5  
    ws.column_dimensions[column].width = adjusted_width

wb.save(output_file)

print(f"Data saved and columns adjusted in {output_file}")
